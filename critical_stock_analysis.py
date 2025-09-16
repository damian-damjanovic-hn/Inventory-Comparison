import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

wb = openpyxl.load_workbook('local_file_dir')
ws = wb.active

headers = [cell.value for cell in ws[1]]
if "StockStatus" not in headers:
    ws.cell(row=1, column=len(headers)+1, value="StockStatus")
    headers.append("StockStatus")
if "Validation" not in headers:
    ws.cell(row=1, column=len(headers)+1, value="Validation")
    headers.append("Validation")

col_qty_wh = headers.index("quantity_warehouse") + 1
col_qty_ec = headers.index("quantity_ecommerce") + 1
col_stock_status = headers.index("StockStatus") + 1
col_validation = headers.index("Validation") + 1

red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

for row in range(2, ws.max_row + 1):
    qty_wh = ws.cell(row=row, column=col_qty_wh).value
    qty_ec = ws.cell(row=row, column=col_qty_ec).value

    if qty_wh is None or qty_ec is None:
        status = "Missing data"
    elif qty_wh <= 0 and qty_ec <= 0:
        status = "Out of stock in both warehouse and ecommerce"
    elif qty_wh <= 0:
        status = "Stock available in ecommerce only"
        ws.cell(row=row, column=col_stock_status).fill = red_fill
    elif qty_ec <= 0:
        status = "Stock available in warehouse only"
        ws.cell(row=row, column=col_stock_status).fill = red_fill
    else:
        status = "Stock available in both warehouse and ecommerce"

    ws.cell(row=row, column=col_stock_status, value=status)

    if qty_wh < 0 or qty_ec < 0:
        validation = "Check negative stock"
        ws.cell(row=row, column=col_validation).fill = red_fill
    else:
        validation = "Stock Qty Data Type Validation PASS"

    ws.cell(row=row, column=col_validation, value=validation)

if "PivotTables" in wb.sheetnames:
    pivot_ws = wb["PivotTables"]
    wb.remove(pivot_ws)

pivot_ws = wb.create_sheet("PivotTables")

from collections import Counter
status_counter = Counter()
validation_counter = Counter()

for row in range(2, ws.max_row + 1):
    status = ws.cell(row=row, column=col_stock_status).value
    validation = ws.cell(row=row, column=col_validation).value
    status_counter[status] += 1
    validation_counter[validation] += 1

pivot_ws.cell(row=1, column=1, value="StockStatus")
pivot_ws.cell(row=1, column=2, value="Count")
row_idx = 2
for key, count in status_counter.items():
    pivot_ws.cell(row=row_idx, column=1, value=key)
    pivot_ws.cell(row=row_idx, column=2, value=count)
    if "only" in key:
        pivot_ws.cell(row=row_idx, column=1).fill = red_fill
    row_idx += 1

pivot_ws.cell(row=row_idx + 1, column=1, value="Validation")
pivot_ws.cell(row=row_idx + 1, column=2, value="Count")
row_idx += 2
for key, count in validation_counter.items():
    pivot_ws.cell(row=row_idx, column=1, value=key)
    pivot_ws.cell(row=row_idx, column=2, value=count)
    if "Check" in key:
        pivot_ws.cell(row=row_idx, column=1).fill = red_fill
    row_idx += 1

wb.save("Updated_Stock_Analysis.xlsx")
print("Macro applied and pivot tables updated in 'Updated_Stock_Analysis.xlsx'.")
